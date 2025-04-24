import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import os


class ExcelGeneration:
    def __init__(self, data, file_name=None):
        """
        :param data: Данные для экспорта (список словарей или DataFrame)
        :param file_name: Необязательно - если нужно сохранить на диск
        """
        self.file_name = file_name
        self.data = data
        self.in_memory = file_name is None  # Режим работы в памяти

    def generate(self):
        """Генерирует Excel и возвращает файл в нужном формате"""
        if self.in_memory:
            return self.__generate_in_memory()
        else:
            self.__generate_to_file()
            return self.file_name

    def __generate_to_file(self):
        """Создает файл на диске"""
        df = pd.DataFrame(self.data)
        df.to_excel(self.file_name, index=False)
        self.__apply_styles(self.file_name)
        print(f"Файл сохранён: {self.file_name}")

    def __generate_in_memory(self):
        """Создает файл в памяти и возвращает BytesIO"""
        output = BytesIO()

        # Сначала сохраняем в временный файл для применения стилей
        temp_file = "temp_export.xlsx"
        df = pd.DataFrame(self.data)
        df.to_excel(temp_file, index=False)
        self.__apply_styles(temp_file)

        # Читаем обратно и пишем в BytesIO
        with open(temp_file, 'rb') as f:
            output.write(f.read())

        os.remove(temp_file)  # Удаляем временный файл
        output.seek(0)
        return output

    def __apply_styles(self, file_path):
        """Применяет стили к файлу"""
        wb = load_workbook(file_path)
        ws = wb.active

        # Настройки стилей
        header_font = Font(bold=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(vertical="top", wrap_text=True)

        # Ширины колонок
        default_widths = {
            1: 18,  # 'Дата публикации'
            2: 40,  # 'Заголовок'
            3: 50,  # 'Краткая суть'
            4: 20,  # 'Источник'
            5: 50,  # 'Ссылка'
        }

        # Применяем стили к заголовкам
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(
                col_idx)].width = default_widths.get(col_idx, 20)

        # Применяем стили к данным
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_align

        wb.save(file_path)
